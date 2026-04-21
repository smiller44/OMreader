"""
Parse the broker / analyst financial workbook (e.g. Eastdil Secured format).
Extracts: unit mix, stabilized projections, tax abatement details.
All values are optional — returns partial dict on missing sheets.
"""
import io
import re

import openpyxl


def _num(v) -> float | None:
    try:
        return float(v) if v is not None else None
    except Exception:
        return None


def _pct_str(v) -> str | None:
    n = _num(v)
    return f"{n * 100:.1f}%" if n is not None else None


def _dollar_str(v) -> str | None:
    n = _num(v)
    if n is None:
        return None
    if abs(n) >= 1_000_000:
        return f"${n / 1_000_000:.2f}M"
    return f"${n:,.0f}"


# ── Rent Roll ─────────────────────────────────────────────────────────────────

def _parse_rent_roll(ws) -> dict:
    """
    Parses a Rent Roll Summary sheet.
    Aggregates sub-types into Studio / 1BR / 2BR / 3BR buckets.
    Returns: units (int), avg_sf, avg_market_rent, avg_leased_rent, unit_mix list.
    """
    buckets: dict[str, dict] = {}
    total_row = None

    for r in range(1, ws.max_row + 1):
        unit_type = ws.cell(r, 3).value
        description = ws.cell(r, 4).value
        count = ws.cell(r, 5).value
        avg_sf = ws.cell(r, 8).value
        market_rent = ws.cell(r, 9).value
        leased_rent = ws.cell(r, 11).value
        proforma_rent = ws.cell(r, 13).value

        if not unit_type or not isinstance(count, (int, float)):
            if description == "Total / Average" and isinstance(count, (int, float)):
                total_row = {
                    "units": int(count),
                    "avg_sf": avg_sf,
                    "avg_market_rent": market_rent,
                    "avg_leased_rent": leased_rent,
                    "avg_proforma_rent": proforma_rent,
                }
            continue

        if str(unit_type).startswith("Total") or description == "Total / Average":
            continue

        desc = str(description).strip()
        if re.search(r"Studio|Stu", desc, re.I):
            bucket = "Studio"
        elif re.search(r"Two Bed|2BR|2 Bed", desc, re.I):
            bucket = "2BR"
        elif re.search(r"Three Bed|3BR|3 Bed", desc, re.I):
            bucket = "3BR"
        elif re.search(r"One Bed|1BR|1 Bed", desc, re.I):
            bucket = "1BR"
        else:
            bucket = desc[:10]

        if bucket not in buckets:
            buckets[bucket] = {"count": 0, "sf_sum": 0.0, "market_sum": 0.0,
                               "leased_sum": 0.0, "proforma_sum": 0.0}
        n = int(count)
        buckets[bucket]["count"]        += n
        buckets[bucket]["sf_sum"]       += float(avg_sf or 0) * n
        buckets[bucket]["market_sum"]   += float(market_rent or 0) * n
        buckets[bucket]["leased_sum"]   += float(leased_rent or 0) * n
        buckets[bucket]["proforma_sum"] += float(proforma_rent or 0) * n

    order = ["Studio", "1BR", "2BR", "3BR"]
    unit_mix = []
    for b in order + [k for k in buckets if k not in order]:
        if b not in buckets:
            continue
        d = buckets[b]
        n = d["count"]
        unit_mix.append({
            "type":           b,
            "count":          n,
            "avg_sf":         round(d["sf_sum"] / n) if n else 0,
            "market_rent":    round(d["market_sum"] / n) if n else 0,
            "leased_rent":    round(d["leased_sum"] / n) if n else 0,
            "proforma_rent":  round(d["proforma_sum"] / n) if n else 0,
        })

    result = {"unit_mix": unit_mix}
    if total_row:
        result.update(total_row)
    return result


# ── Financial Dashboard ───────────────────────────────────────────────────────

def _parse_dashboard(ws) -> dict:
    out = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 3).value
        val   = ws.cell(r, 5).value
        if not label:
            continue
        label_s = str(label).strip().lower()

        if "market rent growth" in label_s:
            out["market_rent_growth"] = _num(val)
        elif "loss to lease" in label_s:
            out["loss_to_lease_assumption"] = _num(val)
        elif "vacancy factor" in label_s:
            out["vacancy_assumption"] = _num(val)
        elif "concession" in label_s:
            out["concession_assumption"] = _num(val)
        elif "expense inflation" in label_s:
            out["expense_growth"] = _num(val)
        elif "bad debt" in label_s:
            out["bad_debt_assumption"] = _num(val)
        elif "management fee" in label_s:
            out["mgmt_fee_pct"] = _num(val)
        elif "reserves" in label_s and "per unit" in label_s:
            out["reserves_per_unit"] = _num(val)
    return out


# ── Stabilized Untrended ─────────────────────────────────────────────────────

def _parse_stabilized(ws) -> dict:
    out = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 3).value
        val   = ws.cell(r, 5).value
        if not label:
            continue
        label_s = str(label).strip().lower()

        if "effective gross" in label_s or "effective rental income" in label_s:
            if "egi" not in out:
                out["stab_egi_raw"] = _num(val)
        elif "net operating income" in label_s:
            if "stab_noi_raw" not in out:
                out["stab_noi_raw"] = _num(val)
        elif "total expenses" in label_s:
            out["stab_opex_raw"] = _num(val)
        elif label_s == "vacancy":
            out["stab_vacancy_raw"] = _num(val)
        elif "total projected market rents" in label_s or "current market rents" in label_s:
            if "stab_gpr_raw" not in out:
                out["stab_gpr_raw"] = _num(val)

    egi  = out.get("stab_egi_raw")
    noi  = out.get("stab_noi_raw")
    opex = out.get("stab_opex_raw")

    result = {
        "stab_label":   "Stabilized Untrended",
        "stab_egi":     _dollar_str(egi),
        "stab_noi":     _dollar_str(noi),
        "stab_opex":    _dollar_str(opex),
        "stab_opex_pct": f"{abs(opex)/egi*100:.1f}%" if (egi and opex) else None,
        "stab_noi_margin": f"{noi/egi*100:.1f}%" if (egi and noi) else None,
    }
    return result


# ── Tax Summary ───────────────────────────────────────────────────────────────

def _parse_tax(ws) -> dict:
    out = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 3).value
        val3  = ws.cell(r, 4).value  # typically year-1 value
        if not label:
            continue
        label_s = str(label).strip().lower()

        if "assessment" in label_s and "re-assess" not in label_s:
            out["tax_assessment"] = _num(val3)
        elif "total net taxes due" in label_s or "total taxes due" == label_s.rstrip():
            if "tax_net_year1" not in out:
                out["tax_net_year1"] = _num(val3)
        elif "total gross taxes due" in label_s:
            out["tax_gross_year1"] = _num(val3)
        elif "abatement" in label_s and "%" in label_s:
            out["abatement_pct_y1"] = _num(val3)
        elif "savings from abatement" in label_s:
            out["abatement_savings_y1"] = _num(val3)
        elif "npv at acquisition" in label_s:
            out["abatement_npv"] = _num(val3)
        elif label_s.startswith("y1 taxes") or "abated" in label_s:
            pass
        elif len(str(label)) > 50 and "brownfield" in str(label).lower():
            out["tax_abatement_desc"] = str(label).strip()

    notes_parts = []
    if out.get("tax_net_year1") and out.get("tax_gross_year1"):
        net  = out["tax_net_year1"]
        gross= out["tax_gross_year1"]
        notes_parts.append(f"Y1 taxes: ${net:,.0f} abated / ${gross:,.0f} unabated.")
    if out.get("abatement_npv"):
        notes_parts.append(f"NPV of abatement: ${out['abatement_npv']:,.0f}.")
    if out.get("tax_abatement_desc"):
        notes_parts.append(out["tax_abatement_desc"])

    out["tax_notes"] = " ".join(notes_parts) if notes_parts else None
    return out


# ── Main entry point ─────────────────────────────────────────────────────────

def parse_financial_workbook(file_bytes: bytes) -> dict:
    """
    Parse an ES-style financial workbook.
    Returns a merged dict ready to overlay onto the deal data dict.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}

    sheet_map = {s.lower(): s for s in wb.sheetnames}

    if "rent roll summary" in sheet_map:
        result.update(_parse_rent_roll(wb[sheet_map["rent roll summary"]]))

    if "financial dashboard" in sheet_map:
        result.update(_parse_dashboard(wb[sheet_map["financial dashboard"]]))

    if "stabilized untrended" in sheet_map:
        result.update(_parse_stabilized(wb[sheet_map["stabilized untrended"]]))

    if "tax summary" in sheet_map:
        result.update(_parse_tax(wb[sheet_map["tax summary"]]))

    # Format for 1-pager
    if result.get("avg_market_rent"):
        result["pro_forma_rent"] = f"${result['avg_market_rent']:,.0f}/mo"
    if result.get("avg_leased_rent"):
        result["in_place_rent"] = f"${result['avg_leased_rent']:,.0f}/mo"
    if result.get("units"):
        result["units"] = str(int(result["units"]))
    if result.get("avg_sf"):
        result["avg_sf"] = f"{int(result['avg_sf'])} SF"

    # unit_mix for 1-pager: keep type + count only
    if result.get("unit_mix"):
        result["unit_mix_detail"] = result["unit_mix"]
        result["unit_mix"] = [{"type": u["type"], "count": u["count"]}
                               for u in result["unit_mix"]]

    return result
