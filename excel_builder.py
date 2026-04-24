"""
Build a pre-filled QuickVal Excel workbook from parsed deal data.
Writes into a copy of templates/quickval_template.xlsx.
"""
import io
import os
from copy import copy
from datetime import datetime, date

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

_WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")

_TEMPLATE = os.path.join(os.path.dirname(__file__), "templates", "quickval_template.xlsx")


def _safe_float(v):
    try:
        return float(v.replace("$", "").replace(",", "").replace("%", "")) if isinstance(v, str) else float(v)
    except Exception:
        return None


def _write_cell(ws, row: int, col: int, value):
    """Write a value and clear any blue input-cell highlight to white."""
    if value is None:
        return
    cell = ws.cell(row, col)
    cell.value = value
    cell.fill = _WHITE_FILL


def _fill_proforma_overview(ws, data: dict, whisper: str = "", mkt: dict | None = None):
    """Fill the Property Overview and Purchase Summary section (col C)."""
    # Row → (col C value)
    fields = {
        5:  data.get("deal_name"),
        6:  data.get("address"),
        7:  data.get("city_state"),
        8:  data.get("submarket"),
        11: data.get("year_built"),
        12: _safe_float(data.get("units")),
        13: _safe_float(str(data.get("avg_sf", "")).replace(" SF", "")),
        17: data.get("broker"),
    }
    for row, val in fields.items():
        _write_cell(ws, row, 3, val)

    # Purchase price
    price_str = whisper or data.get("purchase_price")
    price = _safe_float(price_str)
    if price:
        _write_cell(ws, 21, 3, price)   # Whisper Price
        _write_cell(ws, 23, 3, price)   # Purchase Price

    # Exit cap
    exit_cap = _safe_float(data.get("exit_cap", "").replace("%", "")) if data.get("exit_cap") else None
    if exit_cap:
        _write_cell(ws, 29, 3, exit_cap / 100 if exit_cap > 1 else exit_cap)

    # Proforma assumptions — use Mesirow market forecast if available, else defaults
    market_growth = mkt.get("market_5yr_avg") if mkt else None
    sub_growth    = mkt.get("sub_5yr_avg")    if mkt else None
    _write_cell(ws, 35, 3, market_growth if market_growth else 0.03)  # Market Rent Growth
    _write_cell(ws, 36, 3, sub_growth    if sub_growth    else 0.02)  # MTM Growth (submarket proxy)
    _write_cell(ws, 40, 3, 0.02)   # Expense Growth
    _write_cell(ws, 41, 3, 0.02)   # Insurance Growth


def _fill_t12_intake(ws_intake, t12_parsed: dict):
    """
    Write every individual T12 line item into the T12 Intake sheet.
    Col B = source account code
    Col C = source line item description
    Cols D–O = 12 monthly values
    Col P = row total
    Col Q = Mesirow COA label (must match T12 Clean SUMIF criteria exactly)

    Row 3 col O (col 3+n_months) = anchor date for the backwards date-chain.
    """
    months     = t12_parsed.get("months", [])
    n_months   = t12_parsed.get("n_months", len(months))
    line_items = t12_parsed.get("line_items", [])

    # Set anchor date on row 3 — always col O (15), the template's fixed last-month column
    if months:
        last_month_str = months[-1]
        try:
            import calendar
            anchor_dt = datetime.strptime(last_month_str, "%b %Y")
            last_day  = calendar.monthrange(anchor_dt.year, anchor_dt.month)[1]
            cell = ws_intake.cell(3, 15)
            cell.value = date(anchor_dt.year, anchor_dt.month, last_day)
            cell.number_format = "MMM-YY"
        except Exception:
            pass

    # Write one row per source line item starting at row 4
    COL_ACCT    = 2   # B — source account code
    COL_NAME    = 3   # C — source description
    COL_FIRST   = 4   # D — first month value
    COL_TOTAL   = 16  # P — hardcoded; template SUMIF always looks at col P
    COL_COA     = 17  # Q — Mesirow COA label for SUMIF

    data_row = 4
    for item in line_items:
        monthly = item["monthly"]
        if not any(v != 0 for v in monthly):
            continue

        ws_intake.cell(data_row, COL_ACCT).value  = item["acct"]
        ws_intake.cell(data_row, COL_NAME).value  = item["name"]
        for i, val in enumerate(monthly[:n_months]):
            ws_intake.cell(data_row, COL_FIRST + i).value = round(val, 2)
        ws_intake.cell(data_row, COL_TOTAL).value = round(item["total"], 2)
        ws_intake.cell(data_row, COL_COA).value   = item["coa_label"]
        data_row += 1

    # Write a visible "Net Operating Income" summary row after all line items
    noi_row = data_row
    ws_intake.cell(noi_row, COL_NAME).value = "Net Operating Income"
    # Use seller's reported NOI if captured; fall back to our Python-computed NOI
    noi_value = t12_parsed.get("reported_noi") or t12_parsed.get("_noi")
    if noi_value is not None:
        ws_intake.cell(noi_row, COL_TOTAL).value = round(noi_value, 2)

    # S40 references the NOI row formulaically so it updates if the row is edited
    ws_intake.cell(40, 19).value = f"=P{noi_row}"


def _fill_ret_schedule(ws, data: dict, tax_data: dict | None):
    """
    Fill the Real Estate Tax Schedule (cols V–AF, rows 21–41).

    Col 22 = V  (re-assessment year, current assessed value, millage, etc.)
    Col 23 = W  (W29: also set to current assessed value as starting market value)

    Key inputs:
      V22 (row 22, col 22): Re-Assessment Year — year after closing
      V29 (row 29, col 22): Full Market Value (pre-sale assessed value from tax bill)
      W29 (row 29, col 23): Same — formula chain starts from W29
      V35 (row 35, col 22): Millage Rate (implied from tax bill, else default)
      V38 (row 38, col 22): Non Ad-Valorem Tax (noxious weed, conservation, etc.)
    """
    # Re-assessment year: CFO date year + 1, or current year + 1
    cfo = ws.cell(10, 3).value  # C10 = Initial CFO Date (already in sheet)
    try:
        if isinstance(cfo, (datetime, date)):
            reassess_year = cfo.year + 1
        else:
            reassess_year = datetime.now().year + 1
    except Exception:
        reassess_year = datetime.now().year + 1
    _write_cell(ws, 22, 22, reassess_year)

    if tax_data:
        assessed = tax_data.get("tax_assessment")
        millage  = tax_data.get("implied_millage")
        non_adv  = tax_data.get("non_adv_tax")

        # Current assessed value → Full Market Value pre-sale (V29 and W29)
        if assessed:
            _write_cell(ws, 29, 22, assessed)
            _write_cell(ws, 29, 23, assessed)

        # Implied millage rate
        if millage:
            _write_cell(ws, 35, 22, millage)

        # Non ad-valorem tax (noxious weed, conservation, etc.)
        if non_adv is not None:
            _write_cell(ws, 38, 22, non_adv)


def build_excel(data: dict, t12_parsed=None, whisper: str = "",
                market_data: dict | None = None,
                tax_data: dict | None = None) -> bytes:
    """
    Build a pre-filled QuickVal workbook.
    - data: merged deal dict (from OM + financial workbook + manual)
    - t12_parsed: output of t12_parser.parse_t12()
    - whisper: optional whisper price string
    - tax_data: aggregated output of tax_parser.aggregate_tax_bills()
    Returns raw bytes of the .xlsx file.
    """
    wb = openpyxl.load_workbook(_TEMPLATE)

    # Fill QuickVal Proforma overview
    ws_pf = wb["QuickVal Proforma"]
    _fill_proforma_overview(ws_pf, data, whisper, market_data)

    # Fill RET schedule
    _fill_ret_schedule(ws_pf, data, tax_data)

    # Fill T12 Intake with monthly data
    if t12_parsed:
        ws_intake = wb["T12 Intake"]
        _fill_t12_intake(ws_intake, t12_parsed)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
