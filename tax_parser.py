"""
Parse property tax bill(s).
Accepts PDF (county assessor / King County website export) or Excel.
Returns a dict of tax figures for the RET schedule in the QuickVal model.

Multiple bills for different parcels should be parsed individually and
aggregated by the caller (sum assessed values, sum annual taxes).
"""
import io
import re


def _num(s) -> float | None:
    try:
        return float(str(s).replace(",", "").replace("$", "").strip())
    except Exception:
        return None


def _first_dollar(text: str, pattern: str) -> float | None:
    """Return the first dollar-amount match for a label pattern."""
    m = re.search(pattern + r"[^\d$]*\$?([\d,]+\.?\d*)", text, re.I)
    return _num(m.group(1)) if m else None


def parse_tax_bill(file_bytes: bytes, filename: str = "") -> dict:
    if filename.lower().endswith(".pdf"):
        return _parse_pdf(file_bytes)
    return _parse_excel(file_bytes)


def aggregate_tax_bills(bills: list[dict]) -> dict:
    """
    Combine multiple parcel tax dicts into a single dict by summing
    assessed values and annual taxes, then recomputing implied millage.
    """
    if not bills:
        return {}
    if len(bills) == 1:
        return bills[0]

    out: dict = {}
    for key in ("tax_assessment", "land_value", "improvement_value",
                "tax_annual", "non_adv_tax"):
        vals = [b[key] for b in bills if b.get(key) is not None]
        if vals:
            out[key] = sum(vals)

    # Recompute implied millage from aggregated values
    if out.get("tax_annual") and out.get("tax_assessment") and out["tax_assessment"] > 0:
        out["implied_millage"] = out["tax_annual"] / out["tax_assessment"]

    # Keep first-parcel metadata
    first = bills[0]
    for key in ("tax_year", "levy_code", "tax_notes"):
        if first.get(key):
            out[key] = first[key]

    out["parcel_count"] = len(bills)
    out["tax_notes"] = _build_notes(out)
    return out


# ── PDF ───────────────────────────────────────────────────────────────────────

def _parse_pdf(file_bytes: bytes) -> dict:
    import pdfplumber

    text = ""
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text += (page.extract_text() or "") + "\n"

    # Try parsers in order; fall back to generic patterns
    out = _parse_king_county(text) or _parse_cambridge_ma(text) or _parse_generic(text)
    out["tax_notes"] = _build_notes(out)
    return out


def _parse_king_county(text: str) -> dict | None:
    """
    Parse King County WA website screenshot format.
    Looks for 'Land value' and 'Improvement value' rows with dollar amounts,
    and a 'Tax' charge row (NOT 'Total billed', which includes fees).
    Takes the FIRST (most recent) year's values in each row.
    """
    DOLLAR = r"\$?([\d,]+\.?\d*)"

    land = _first_dollar(text, r"Land\s+value")
    impr = _first_dollar(text, r"Improvement\s+value")
    if land is None or impr is None:
        return None

    assessed = land + impr

    # 'Tax' row — first amount only (most recent year)
    tax_m = re.search(r"(?<!\w)Tax\s+" + DOLLAR, text, re.I)
    annual_tax = _num(tax_m.group(1)) if tax_m else None

    # Non ad-valorem fees: noxious weed, conservation, surface water, etc.
    non_adv = 0.0
    for label in (r"Noxious\s+Weed", r"Conservation", r"Surface\s+Water",
                  r"Flood\s+Control", r"Fire\s+District", r"Library"):
        m = re.search(label + r"\s+" + DOLLAR, text, re.I)
        if m:
            non_adv += _num(m.group(1)) or 0.0

    # Levy code
    levy_m = re.search(r"Levy\s+code\s+(\d+)", text, re.I)
    levy_code = levy_m.group(1) if levy_m else None

    # Tax year — first 4-digit year that looks like a tax year
    year_m = re.search(r"(?:Tax\s+Information|Breakdown\s+by\s+Tax\s+Year)[^\d]*(\d{4})", text, re.I)
    tax_year = year_m.group(1) if year_m else None

    out: dict = {
        "land_value":        land,
        "improvement_value": impr,
        "tax_assessment":    assessed,
        "non_adv_tax":       non_adv if non_adv > 0 else None,
    }
    if annual_tax:
        out["tax_annual"] = annual_tax
    if levy_code:
        out["levy_code"] = levy_code
    if tax_year:
        out["tax_year"] = tax_year

    if annual_tax and assessed > 0:
        out["implied_millage"] = annual_tax / assessed

    return out


def _parse_cambridge_ma(text: str) -> dict | None:
    """
    Parse Cambridge MA / Massachusetts assessor tax bill format.
    Looks for 'Total Taxable Value' (assessed) and 'RES TAX' or tax rate patterns.
    Also handles 'Community Preservation Act' (CPA) surcharge as non ad-valorem.
    Returns None if this doesn't look like a Cambridge MA bill.
    """
    DOLLAR = r"\$?([\d,]+\.?\d*)"

    # Assessed value — Cambridge uses "Total Taxable Value" or "Assessed Value"
    assessed = None
    for pat in [r"Total\s+Taxable\s+Value\s*[:\-]?\s*" + DOLLAR,
                r"Assessed\s+Value\s*[:\-]?\s*" + DOLLAR]:
        m = re.search(pat, text, re.I)
        if m:
            assessed = _num(m.group(1))
            break

    if assessed is None:
        return None

    # Annual tax — "RES TAX" line is the base property tax (exclude CPA surcharge)
    annual_tax = None
    for pat in [r"RES\s+TAX\s*" + DOLLAR,
                r"(?:Real\s+Estate\s+)?Tax\s*[:\-]?\s*" + DOLLAR]:
        m = re.search(pat, text, re.I)
        if m:
            annual_tax = _num(m.group(1))
            break

    # CPA (Community Preservation Act) surcharge → non ad-valorem
    non_adv = None
    m = re.search(r"CPA\s*" + DOLLAR, text, re.I)
    if m:
        non_adv = _num(m.group(1))

    # Tax year
    tax_year = None
    m = re.search(r"FY\s*(\d{4})", text, re.I)
    if m:
        tax_year = m.group(1)

    out: dict = {"tax_assessment": assessed}
    if annual_tax:
        out["tax_annual"] = annual_tax
    if non_adv:
        out["non_adv_tax"] = non_adv
    if tax_year:
        out["tax_year"] = tax_year
    if annual_tax and assessed > 0:
        out["implied_millage"] = annual_tax / assessed

    return out


def _parse_generic(text: str) -> dict:
    """Fallback patterns for traditional county assessor PDF notices."""
    out: dict = {}

    for pat in [
        r"total\s+assessed\s+value[^\d$]*\$?([\d,]+)",
        r"assessed\s+value[^\d$]*\$?([\d,]+)",
    ]:
        m = re.search(pat, text, re.I)
        if m:
            out["tax_assessment"] = _num(m.group(1))
            break

    for pat in [
        r"total\s+gross\s+tax(?:es)?[^\d$]*\$?([\d,]+)",
        r"gross\s+tax(?:es)?\s+(?:due|billed)[^\d$]*\$?([\d,]+)",
        r"total\s+taxes?\s+(?:billed|levied)[^\d$]*\$?([\d,]+)",
        r"total\s+billed[^\d$]*\$?([\d,]+)",
    ]:
        m = re.search(pat, text, re.I)
        if m:
            out["tax_annual"] = _num(m.group(1))
            break

    if out.get("tax_annual") and out.get("tax_assessment") and out["tax_assessment"] > 0:
        out["implied_millage"] = out["tax_annual"] / out["tax_assessment"]

    m = re.search(r"(?:tax\s+year|year)[:\s]+(\d{4})", text, re.I)
    if m:
        out["tax_year"] = m.group(1)

    return out


# ── Excel ─────────────────────────────────────────────────────────────────────

def _parse_excel(file_bytes: bytes) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = next(
        (wb[s] for s in wb.sheetnames if "tax" in s.lower()),
        wb.active,
    )

    out: dict = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 3).value
        val   = ws.cell(r, 4).value
        if not label:
            continue
        ls = str(label).strip().lower()

        if "assessment" in ls and "re-assess" not in ls:
            out["tax_assessment"] = _num(val)
        elif "total net taxes due" in ls or ls.rstrip() == "total taxes due":
            out.setdefault("tax_annual", _num(val))
        elif "total gross taxes due" in ls or "total billed" in ls:
            out["tax_annual"] = _num(val)
        elif "savings from abatement" in ls:
            out["abatement_savings_y1"] = _num(val)
        elif "npv at acquisition" in ls:
            out["abatement_npv"] = _num(val)

    if out.get("tax_annual") and out.get("tax_assessment") and out["tax_assessment"] > 0:
        out["implied_millage"] = out["tax_annual"] / out["tax_assessment"]

    out["tax_notes"] = _build_notes(out)
    return out


# ── Shared ────────────────────────────────────────────────────────────────────

def _build_notes(out: dict) -> str | None:
    parts = []
    year = out.get("tax_year", "")
    prefix = f"{year} " if year else ""

    annual = out.get("tax_annual")
    assessed = out.get("tax_assessment")
    millage = out.get("implied_millage")

    if annual:
        parts.append(f"{prefix}Annual tax: ${annual:,.0f}.")
    if assessed:
        parts.append(f"Assessed value: ${assessed:,.0f}.")
    if millage:
        parts.append(f"Implied millage: {millage*100:.4f}%.")
    if out.get("non_adv_tax"):
        parts.append(f"Non ad-valorem fees: ${out['non_adv_tax']:,.2f}.")
    if out.get("parcel_count", 1) > 1:
        parts.append(f"({out['parcel_count']} parcels aggregated)")

    return " ".join(parts) if parts else None
