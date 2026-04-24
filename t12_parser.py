"""
Parse a property-management T12 export into structured data for QuickVal.
All account-code → COA mapping is handled by Claude (via extra_mappings).
"""
import io
import re
from datetime import datetime, date as date_type

import openpyxl


# COA code → display label (used both for T12 Intake col Q and summary display)
COA_LABELS: dict[str, str] = {
    "mkt":     "Gross Potential Rent - Market Rate",
    "aff":     "Gross Potential Rent - Affordables",
    "ltl":     "Gain / Loss-to-Lease",
    "vac":     "Physical Vacancy",
    "conc":    "Concessions",
    "empmo":   "Employee / Model Units",
    "down":    "Down Units",
    "bd":      "Bad Debt / Write-Offs",
    "pkg":     "Parking",
    "stg":     "Storage",
    "pet":     "Pet Fees / Rent",
    "tv":      "Cable / Internet",
    "oinc":    "Other Income Misc.",
    "rubs":    "Utility Reimbursements",
    "cominc":  "Commercial Income",
    "adv":     "Advertising / Marketing",
    "adm":     "Administrative",
    "rm":      "Repairs & Maintenance",
    "cs":      "Contract Services",
    "to":      "Turnover / Make Ready",
    "pay":     "Payroll & Benefits",
    "util":    "Utilities",
    "mgt":     "Management Fee",
    "ins":     "Insurance",
    "ret":     "Real Estate Taxes",
    "hoa":     "HOA",
    "comexp":  "Commercial Expenses",
    "capx":    "Replacement Reserves",
    "nrcapx":  "Non-Recurring CapEx",
    "amcapx":  "Amenity / Common Area CapEx",
    "intcapx": "Unit Interior CapEx",
    "tilc":    "TI&LC Expenditures",
    "nai":     "N/A Income / Entity Income",
    "nae":     "N/A Expenses / Entity Expenses",
}

# Rich descriptions for the Claude classification prompt
COA_DESCRIPTIONS: dict[str, str] = {
    "mkt":     "Gross potential rent for market-rate units — scheduled rent at 100% occupancy",
    "aff":     "Gross potential rent for affordable, subsidized, or income-restricted units",
    "ltl":     "Gain or loss-to-lease — difference between market rent and actual contracted/leased rent",
    "vac":     "Physical vacancy loss — rent lost due to unoccupied units",
    "conc":    "Concessions — free rent, move-in specials, look & lease discounts, one-time incentives",
    "empmo":   "Employee units, model units, leasing office units, guest suites used by staff",
    "down":    "Down units — offline units under renovation or held out of service",
    "bd":      "Bad debt, write-offs, collections, uncollected rent charged off",
    "pkg":     "Parking income — carport, garage, covered, reserved, or surface lot fees",
    "stg":     "Storage unit rent, storage locker fees",
    "pet":     "Pet fees (non-refundable), pet rent, pet deposits kept as income",
    "tv":      "Bulk cable TV, internet, Wi-Fi, media services billed to residents",
    "oinc":    (
        "Other miscellaneous income: admin fees, application fees, late charges, MTM premiums, "
        "NSF fees, transfer fees, smart home fees, renter's insurance income, damages collected, "
        "lease cancellation fees, key/lock fees, club room rental, community fees, "
        "guest suite income, interest income, vendor rebates, access gate remote fees"
    ),
    "rubs":    "Utility reimbursements billed back to tenants — water/sewer rebill, trash rebill, pest control rebill, utility RUBS income",
    "cominc":  "Commercial or retail tenant rental income, ground floor retail rents",
    "adv":     (
        "Advertising & marketing: ILS listings, SEM/PPC campaigns, property website, "
        "social media, reputation management, signage, print/publications, "
        "outreach marketing, broker/locator referral fees, resident events, "
        "prospect refreshments, resident retention programs, tour experience"
    ),
    "adm":     (
        "Administrative: office supplies, postage, cell phones, telephone, internet access, "
        "software licenses, bank charges, legal fees, eviction fees, computer expense, "
        "training/seminars, employee recruitment, uniform rental, licenses/fees/permits, "
        "payment processing fees, business automation, revenue management software"
    ),
    "rm":      (
        "Repairs & maintenance: appliance repairs, HVAC, plumbing, electrical, "
        "building exterior/interior, common area repairs, painting, lighting, locks/keys, "
        "maintenance supplies, preventative maintenance, safety/fire, small tools, "
        "sprinklers, dog park, pool/spa, gym/fitness, business center, club room, amenities"
    ),
    "cs":      (
        "Contract services: landscaping, janitorial, pest control, elevator contract, "
        "fire alarm/suppression, fire protection, patrol/security/courtesy officer, "
        "snow removal, trash removal, door-to-door trash, pool & spa contract, "
        "access gate contract, cable TV contract, music/video/TV service, equipment contracts"
    ),
    "to":      "Turnover & make-ready: unit cleaning, housekeeper, paint contractor, painting supplies, blinds/drapes, carpet, touch-up between tenants",
    "pay":     "Payroll & benefits: all property staff salaries (manager, leasing, maintenance, assistant roles), bonuses, 401k, group health insurance, employee burden/taxes",
    "util":    "Utilities paid by owner: electricity for common areas, vacant units, models; gas; water/sewer (owner-paid); utility billing service/RUBS admin fees",
    "mgt":     "Property management fees, asset management fees, management company charges",
    "ins":     "Property insurance, casualty insurance, liability insurance, umbrella policy premiums",
    "ret":     "Real estate taxes, ad valorem property taxes, personal property taxes, tax assessments",
    "hoa":     "HOA dues, condo association fees, master association fees",
    "comexp":  "Commercial or retail tenant expenses, ground floor retail operating expenses",
    "capx":    "Replacement reserves, capital expenditure reserves, recurring capex set-asides",
    "nrcapx":  "Non-recurring capital expenditures — large one-time repairs or replacements",
    "amcapx":  "Amenity or common area capital improvements — pool renovation, gym upgrade, lobby remodel",
    "intcapx": "Unit interior capital improvements — unit renovation, value-add upgrades, appliance replacements",
    "tilc":    "Tenant improvement allowances, leasing commissions for commercial tenants",
    "nai":     "Non-operating or entity-level income — owner distributions received, interest on reserves, items that don't belong in NOI",
    "nae":     "Non-operating or entity-level expenses — interest expense, depreciation, amortization, owner distributions, items that don't belong in NOI",
}

# Hardcoded prefix → COA code for common Yardi/MRI/RealPage account ranges.
# Claude is only called for prefixes NOT found here.
_YARDI_ACCT_MAP: dict[str, str] = {
    # ── Rental Income ──────────────────────────────────────────────────────
    "3110": "mkt",   # Gross Potential Rent (market)
    "3115": "mkt",
    "3120": "aff",   # Affordable / subsidized GPR
    "3125": "aff",
    "3150": "mkt",   # Market Rent
    "3155": "mkt",
    "3160": "ltl",   # Gain / Loss-to-Lease
    "3165": "ltl",
    "3170": "ltl",
    "3175": "ltl",
    "3180": "vac",   # Vacancy Loss
    "3185": "vac",
    "3190": "vac",
    "3200": "empmo", # Employee / Model Units
    "3205": "empmo",
    "3210": "down",  # Down / Offline Units
    "3215": "down",
    "3270": "bd",    # Bad Debt / Write-Offs
    "3271": "bd",
    "3275": "bd",
    "3280": "bd",
    "3285": "bd",    # Bad Debt Recovery (nets against bd)
    "3290": "bd",
    "3295": "bd",
    "3300": "conc",  # Concessions / Free Rent
    "3305": "conc",
    "3310": "conc",
    "3315": "conc",
    "3320": "conc",
    "3325": "conc",
    "3330": "conc",
    # ── Other Income ────────────────────────────────────────────────────────
    "3500": "oinc",  # Other Income (generic)
    "3510": "oinc",
    "3512": "oinc",  # Non-Refundable Admin Fees
    "3513": "oinc",
    "3514": "oinc",
    "3515": "oinc",
    "3516": "pet",   # Non-Refundable Pet Fees
    "3517": "pet",
    "3518": "pet",
    "3520": "oinc",  # Application Fees
    "3522": "oinc",
    "3524": "oinc",  # Late Fees
    "3526": "oinc",  # NSF Fees
    "3528": "oinc",  # Termination Fees / Lease Break
    "3530": "oinc",  # Month-to-Month Premiums
    "3535": "oinc",
    "3540": "oinc",
    "3542": "stg",   # Storage
    "3544": "pkg",   # Parking - Carport
    "3545": "pkg",
    "3546": "pkg",   # Parking - Garage
    "3547": "pkg",
    "3548": "pkg",
    "3549": "pkg",
    "3550": "tv",    # Cable / Internet / Bulk TV
    "3551": "tv",
    "3552": "tv",
    "3553": "tv",
    "3554": "oinc",  # Interest Income (operating, not entity-level)
    "3555": "oinc",
    "3556": "oinc",
    "3558": "oinc",  # Other Income
    "3560": "oinc",
    "3562": "pet",   # Pet Rent
    "3563": "pet",   # Pet Rent
    "3564": "pet",
    "3565": "pet",
    "3570": "oinc",
    "3575": "oinc",
    "3578": "oinc",
    "3580": "bd",    # Bad Debt Recovery (other income side)
    "3582": "oinc",  # Damages Collected
    "3585": "oinc",
    "3590": "rubs",  # Utility Reimbursements - Electricity
    "3591": "rubs",
    "3592": "rubs",  # Billback - Sewer
    "3593": "rubs",  # Utility Billing Fees
    "3594": "rubs",  # Billback - Waste / Trash
    "3595": "rubs",
    "3596": "rubs",
    "3597": "rubs",  # Billback - Water/Sewer
    "3598": "rubs",
    "3599": "rubs",
    "3600": "oinc",
    "3610": "oinc",
    "3620": "oinc",
    "3630": "oinc",
    "3639": "oinc",  # Renter's Insurance Income
    "3640": "oinc",  # Renter's Insurance Program
    "3650": "oinc",
    "3660": "cominc", # Commercial / Retail Income
    "3665": "cominc",
    "3670": "cominc",
    "3680": "oinc",
    "3690": "oinc",
    # ── Payroll ─────────────────────────────────────────────────────────────
    "4010": "pay",
    "4011": "pay",
    "4012": "pay",   # Manager Salary
    "4013": "pay",
    "4014": "pay",
    "4015": "pay",
    "4016": "pay",
    "4017": "pay",
    "4018": "pay",   # Leasing Salary
    "4019": "pay",
    "4020": "pay",
    "4021": "pay",   # Bonuses / Benefits / Medical
    "4022": "pay",   # FICA / FUTA / Benefits
    "4023": "pay",   # SUTA
    "4024": "pay",
    "4025": "pay",
    "4026": "pay",   # Workers Comp
    "4027": "pay",   # 401k / Temp Help
    "4028": "pay",
    "4029": "pay",
    "4030": "pay",   # Maintenance Payroll
    "4031": "pay",   # Maintenance Tech / Rent Allowance
    "4032": "pay",   # R&M Bonuses / FICA / FUTA
    "4033": "pay",   # R&M SUTA
    "4034": "pay",
    "4035": "pay",   # Workers Comp / Medical Insurance
    "4036": "pay",   # 401k
    "4037": "pay",
    "4038": "pay",
    "4039": "pay",
    "4040": "pay",
    "4045": "pay",
    "4050": "pay",
    "4055": "pay",
    "4060": "pay",
    "4065": "pay",
    "4070": "pay",
    "4075": "pay",
    "4080": "pay",
    "4085": "pay",
    "4090": "pay",
    "4095": "pay",
    # ── Administrative ──────────────────────────────────────────────────────
    "4100": "adm",
    "4105": "adm",
    "4110": "adm",
    "4115": "adm",
    "4120": "adm",
    "4125": "adm",
    "4130": "adm",
    "4135": "adm",
    "4140": "adm",
    "4144": "adm",   # Employee Screening
    "4145": "adm",
    "4148": "adm",   # Bank Charges
    "4150": "adm",
    "4155": "adm",
    "4158": "adm",   # Payroll Processing Fees
    "4160": "adm",   # Dues & Subscriptions
    "4161": "adm",   # Permits & Licenses
    "4162": "adm",   # Software Licenses
    "4163": "adm",   # Revenue Management Software
    "4165": "adm",
    "4170": "adm",
    "4175": "adm",
    "4180": "adm",
    "4185": "adm",
    "4190": "adm",
    "4192": "adm",   # Legal / Evictions
    "4195": "adm",
    "4200": "adm",
    "4205": "adm",
    "4208": "adm",   # Training
    "4210": "adm",
    "4213": "hoa",   # HOA Dues ← non-variable, not adm
    "4215": "adm",
    "4220": "adm",
    "4223": "mgt",   # Management Fee ← non-variable
    "4225": "adm",
    "4228": "adm",   # Office Refreshments
    "4230": "adm",
    "4235": "adm",
    "4240": "adm",
    "4245": "adm",
    "4248": "adm",
    "4250": "adm",   # Office Supplies
    "4255": "adm",
    "4256": "adm",   # Postage
    "4258": "adm",
    "4260": "adm",
    "4264": "adm",   # Computer / Modem
    "4265": "adm",
    "4267": "adm",   # Pagers
    "4268": "adm",   # Office Telephone
    "4269": "adm",   # Internet
    "4270": "adm",   # Copier
    "4275": "adm",
    "4278": "adm",
    "4280": "adm",   # Meals / Entertainment
    "4285": "adm",
    "4290": "adm",
    "4295": "adm",
    # ── Marketing / Advertising ─────────────────────────────────────────────
    "4300": "adv",
    "4305": "adv",
    "4310": "adv",
    "4311": "adv",   # ILS / Online Advertising
    "4312": "adv",
    "4315": "adv",
    "4320": "adv",
    "4325": "adv",
    "4330": "adv",
    "4332": "adv",   # Resident Activities / Events
    "4335": "adv",
    "4340": "adv",
    "4344": "adv",   # Photography
    "4345": "adv",
    "4350": "adv",
    "4355": "adv",
    "4360": "adv",
    "4365": "adv",
    "4366": "adv",   # Resident Referrals
    "4368": "adv",
    "4370": "adv",   # Tenant Screening
    "4375": "adv",
    "4380": "adv",
    "4385": "adv",
    "4390": "adv",
    "4395": "adv",   # Marketing Misc
    "4397": "adv",
    "4398": "adv",
    "4399": "adv",
    # ── Repairs & Maintenance ───────────────────────────────────────────────
    "4400": "rm",
    "4405": "rm",
    "4410": "rm",
    "4415": "rm",
    "4420": "rm",
    "4425": "rm",
    "4430": "rm",
    "4431": "rm",    # R&M Supplies
    "4432": "rm",    # Appliance Supplies
    "4433": "rm",
    "4434": "rm",    # Cleaning Supplies
    "4435": "rm",    # Doors / Windows
    "4436": "rm",    # Electrical Supplies
    "4437": "rm",
    "4438": "rm",
    "4439": "rm",    # Fire / Alarm Supplies
    "4440": "rm",    # Glass / Screen
    "4441": "rm",
    "4442": "rm",    # Hardware
    "4443": "rm",
    "4444": "rm",    # HVAC Supplies
    "4445": "rm",
    "4446": "rm",    # Keys / Locks
    "4447": "rm",    # Light Bulbs
    "4448": "rm",    # Paint Supplies
    "4449": "rm",
    "4450": "rm",
    "4454": "rm",    # Plumbing Supplies
    "4455": "rm",
    "4460": "rm",
    "4465": "rm",
    "4470": "rm",
    "4475": "rm",
    "4478": "rm",    # Clubhouse & Amenities
    "4480": "rm",
    "4484": "rm",    # Electrical Services
    "4485": "rm",
    "4487": "rm",    # Golf Cart / Equipment Service
    "4490": "rm",
    "4494": "rm",    # Pest Control
    "4495": "rm",
    "4496": "rm",    # Plumbing Services
    "4498": "rm",
    "4499": "rm",
    # ── Contract Services ───────────────────────────────────────────────────
    "4500": "cs",
    "4505": "cs",
    "4510": "cs",
    "4515": "cs",
    "4520": "cs",
    "4525": "cs",
    "4530": "cs",
    "4535": "cs",
    "4540": "cs",
    "4545": "cs",
    "4550": "cs",
    "4552": "cs",    # Landscape - Annual Contract
    "4555": "cs",
    "4558": "cs",    # Landscape - Seasonal
    "4560": "cs",
    "4565": "cs",
    "4570": "cs",    # Snow Removal
    "4575": "cs",
    "4580": "cs",
    "4585": "cs",
    "4590": "cs",
    "4595": "cs",
    "4600": "cs",
    "4605": "cs",
    "4610": "cs",
    "4615": "cs",
    "4620": "cs",
    "4621": "cs",
    "4622": "cs",
    "4623": "cs",
    "4624": "cs",
    "4625": "cs",
    "4626": "cs",    # Security / Fire Protection
    "4630": "cs",
    "4631": "cs",
    "4632": "cs",    # Alarm Service
    "4635": "cs",
    "4640": "cs",
    "4645": "cs",
    "4650": "cs",
    "4655": "cs",
    "4660": "cs",
    "4665": "cs",
    "4670": "cs",
    "4675": "cs",
    "4680": "cs",
    "4685": "cs",
    "4690": "cs",
    "4695": "cs",
    # ── Turnover / Make Ready ────────────────────────────────────────────────
    "4700": "to",
    "4701": "to",    # Blind Replacement
    "4702": "to",    # Carpet Cleaning
    "4705": "to",
    "4710": "to",
    "4713": "to",    # Contract Cleaning
    "4715": "to",
    "4720": "to",
    "4723": "to",    # Turnover Misc
    "4724": "to",    # Painting / Paint Supplies
    "4725": "to",
    "4730": "to",
    "4735": "to",
    "4740": "to",
    "4745": "to",
    "4750": "to",
    "4755": "to",
    "4760": "to",
    "4765": "to",
    "4770": "to",
    "4775": "to",
    "4780": "to",
    "4785": "to",
    "4790": "to",
    "4795": "to",
    # ── Utilities ───────────────────────────────────────────────────────────
    "4800": "util",
    "4802": "util",  # Electricity - Common Area
    "4803": "util",
    "4804": "util",  # Electricity - Vacant
    "4805": "util",
    "4806": "util",
    "4810": "util",
    "4812": "util",  # Gas - Common Area
    "4815": "util",
    "4817": "util",  # Gas - Vacant
    "4820": "util",  # Sewer
    "4825": "util",
    "4830": "util",
    "4835": "util",
    "4840": "util",
    "4844": "util",  # Waste Removal
    "4845": "util",
    "4850": "util",
    "4852": "util",  # Water - Common Area
    "4855": "util",
    "4860": "util",  # Utility Processing Fees
    "4865": "util",
    "4870": "util",
    "4875": "util",
    "4880": "util",
    "4885": "util",
    "4890": "util",
    "4895": "util",
    # ── Insurance ───────────────────────────────────────────────────────────
    "5000": "ins",
    "5001": "ins",
    "5002": "ins",
    "5003": "ins",
    "5004": "ins",   # Insurance - Expense / Other
    "5005": "ins",
    "5010": "ins",
    "5015": "ins",
    "5020": "ins",
    # ── Real Estate Taxes ────────────────────────────────────────────────────
    "5006": "ret",   # Property Taxes
    "5007": "ret",
    "5008": "ret",
    "5025": "ret",
    "5030": "ret",
    "5035": "ret",
    "5040": "ret",
    "5045": "ret",
    "5050": "ret",
    "5100": "ret",
    "5150": "ret",
    "5200": "ret",
    # ── Management Fee ───────────────────────────────────────────────────────
    "5300": "mgt",
    "5310": "mgt",
    "5320": "mgt",
    "5400": "mgt",
    # ── Capital / Reserves ───────────────────────────────────────────────────
    "6000": "capx",
    "6005": "capx",
    "6010": "capx",
    "6015": "capx",
    "6020": "capx",
    "6050": "nrcapx",
    "6100": "nrcapx",
    "6150": "amcapx",
    "6200": "intcapx",
    # ── Non-Operating (entity-level; excluded from NOI) ───────────────────
    "7000": "nae",   # Interest Expense
    "7100": "nae",   # Depreciation
    "7200": "nae",   # Amortization
    "7300": "nae",
    "7400": "nae",
    "7500": "nae",
    "8000": "nae",
    "8100": "nae",
    "8200": "nai",
    "8500": "nae",
    "9000": "nae",
    "9100": "nai",
}

_NOI_PATTERNS = (
    "net operating income", "net operating cash flow", "operating cash flow",
    "total noi", "property noi", "net income from operations",
    "net operating",
)

# High-level aggregate rows that are always skipped (pure roll-ups we never want).
_ALWAYS_SKIP = frozenset({
    "net operating", "effective gross", "n/a income", "n/a expense",
    "operating cash flow", "total revenue", "total net rental income",
    "total rental income", "total other operating income",
    "total net potential rent", "total non revenue units",
    "total controllable expenses", "total expenses",
    "net rental income", "total net rental",
    # Entrata/Griffis aggregates (EGI-level, not NOI)
    "total operating income", "total income", "net income",
    # Entrata section subtotals (all-caps, no "total" prefix)
    "operating expenses",
})


def _acct_prefix(code: str) -> str:
    return str(code).split("-")[0].strip()


_ACCT_RE = re.compile(r"^\d{4,5}(-\d+)*$")
_ACCT_INLINE_RE = re.compile(r"^(\d{4,5}(?:-\d+)*)\s+(.+)$")


def _extract_acct_name(row: list, name_only: bool = False) -> tuple[str, str]:
    """Return (acct_code, name) handling variable column layouts.

    name_only=True: Entrata/Griffis format where col A = description,
    col B = first month value (no separate account-code column).
    """
    col0 = str(row[0]).strip() if row else ""
    if name_only:
        return "", col0

    col1 = str(row[1]).strip() if len(row) > 1 else ""
    col2 = str(row[2]).strip() if len(row) > 2 else ""

    if _ACCT_RE.match(col0):
        return col0, col1
    if not col0 and _ACCT_RE.match(col1):
        return col1, col2
    m = _ACCT_INLINE_RE.match(col0)
    if m:
        return m.group(1), m.group(2)
    return col0, col1


def _to_float(v) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",", "").strip())
        except ValueError:
            pass
    return 0.0


# ── Excel header detection ─────────────────────────────────────────────────────

_MONTH_STR_RE = re.compile(
    r"(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
    r"['\-.\s,/]*(\d{2,4})?",
    re.I,
)

# MM/DD/YYYY or MM/DD/YY date strings (e.g. CBRE Condensed format "03/31/2025")
_MDY_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$")

_BARE_MONTH_RE = re.compile(
    r"^(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)$",
    re.I,
)

_MONTHS_ORDER = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]


def _month_label(mon_str: str, year_str: str | None, prev_label: str | None) -> str:
    """Build 'Mon YYYY' string, inferring year from sequence if absent."""
    mon3 = mon_str[:3].capitalize()
    if year_str:
        yr = "20" + year_str if len(year_str) == 2 else year_str
        return f"{mon3} {yr}"
    if prev_label:
        try:
            prev_dt = datetime.strptime(prev_label, "%b %Y")
            # Advance one month
            m = prev_dt.month % 12 + 1
            y = prev_dt.year + (1 if prev_dt.month == 12 else 0)
            return datetime(y, m, 1).strftime("%b %Y")
        except Exception:
            pass
    return f"{mon3} ???"


def _detect_header_row(ws) -> tuple[int, int, int, list[str]]:
    for r in range(1, 50):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        dates, total_col = [], None
        for ci, v in enumerate(row, 1):
            if isinstance(v, (datetime, date_type)):
                dt = v if isinstance(v, datetime) else datetime(v.year, v.month, v.day)
                dates.append((ci, dt.strftime("%b %Y")))
            elif isinstance(v, str):
                s = v.strip()
                if "total" in s.lower() and not _MONTH_STR_RE.match(s):
                    total_col = ci
                    continue
                # MM/DD/YYYY (CBRE Condensed format)
                mdy = _MDY_RE.match(s)
                if mdy:
                    mo, day, yr = int(mdy.group(1)), int(mdy.group(2)), mdy.group(3)
                    yr = "20" + yr if len(yr) == 2 else yr
                    try:
                        dt = datetime(int(yr), mo, 1)
                        dates.append((ci, dt.strftime("%b %Y")))
                        continue
                    except ValueError:
                        pass
                m = _MONTH_STR_RE.match(s)
                if m:
                    prev = dates[-1][1] if dates else None
                    dates.append((ci, _month_label(m.group(1), m.group(2), prev)))
                elif _BARE_MONTH_RE.match(s):
                    prev = dates[-1][1] if dates else None
                    dates.append((ci, _month_label(s, None, prev)))
        if len(dates) >= 3:
            if total_col is None:
                total_col = dates[-1][0] + 1
            return r, dates[0][0], total_col, [d[1] for d in dates]
    raise ValueError("Could not detect month header row in T12 file.")


# ── PDF extraction ─────────────────────────────────────────────────────────────

def _pdf_to_rows(file_bytes: bytes) -> list[list]:
    import pdfplumber
    all_rows = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            for tbl in (page.extract_tables() or []):
                for row in tbl:
                    if any(c for c in row if c):
                        all_rows.append([str(c).strip() if c else "" for c in row])
    return all_rows


def _detect_header_row_pdf(rows: list[list]):
    for i, row in enumerate(rows):
        dates, total_col = [], None
        for ci, v in enumerate(row):
            s = str(v).strip() if v else ""
            if s.lower() == "total":
                total_col = ci
                continue
            m = _MONTH_STR_RE.match(s)
            if m:
                prev = dates[-1][1] if dates else None
                dates.append((ci, _month_label(m.group(1), m.group(2), prev)))
            elif _BARE_MONTH_RE.match(s):
                prev = dates[-1][1] if dates else None
                dates.append((ci, _month_label(s, None, prev)))
        if len(dates) >= 3:
            if total_col is None:
                total_col = dates[-1][0] + 1
            return i, dates[0][0], total_col, [d[1] for d in dates]
    raise ValueError("Could not detect month header row in PDF T12.")


# ── Core parse ─────────────────────────────────────────────────────────────────

def _parse_rows(rows: list[list], hdr_idx: int, first_col: int, total_col: int,
                acct_map: dict, name_only: bool = False):
    n_months = total_col - first_col
    line_items, unmapped, coa = [], [], {}
    reported_noi = None

    # Track whether any individual (account-code) data rows have been seen since
    # the last section-header boundary. Used to decide whether "Total X" leaf rows
    # (present in CBRE Condensed and similar formats) should be kept or dropped.
    section_had_accounts = False

    # For name-only (Entrata) format: track income vs expense section so identical
    # item names in different sections get distinct prefixes for Claude classification.
    entrata_section = "unknown"  # "income" | "expense" | "unknown"

    for row in rows[hdr_idx + 1:]:
        if len(row) <= total_col:
            continue

        acct_raw, name_raw = _extract_acct_name(row, name_only)
        # Allow blank acct_raw as long as we have a name (e.g. CBRE blank-code rows)
        if not name_raw:
            continue

        name_lower = name_raw.strip().lower()

        monthly = [_to_float(row[first_col + i] if first_col + i < len(row) else 0) for i in range(n_months)]
        while len(monthly) < 12:
            monthly.append(0.0)
        total = sum(monthly[:n_months])
        has_values = any(v != 0 for v in monthly[:n_months])

        # In name-only format, zero-value rows are section headers (e.g. "  INCOME").
        # Detect income/expense boundary before skipping.
        if name_only and not acct_raw and not has_values:
            nm = name_raw.strip().upper()
            if nm == "INCOME":
                entrata_section = "income"
            elif nm in ("EXPENSES", "EXPENSE"):
                entrata_section = "expense"
            continue

        # ── NOI capture (must happen before skip guards) ──────────────────────
        if reported_noi is None and any(p in name_lower for p in _NOI_PATTERNS):
            if has_values:
                reported_noi = total

        # ── Always-skip aggregates ───────────────────────────────────────────
        if any(p in name_lower for p in _ALWAYS_SKIP):
            # Treat as a section boundary so the next "Total X" starts fresh
            if has_values:
                section_had_accounts = False
            continue

        # ── Section headers: same text in both name columns, no values ───────
        # In CBRE Condensed format, section headers repeat col A in col B.
        col0 = str(row[0]).strip() if row else ""
        col1 = str(row[1]).strip() if len(row) > 1 else ""
        is_same_col = col0 == col1
        if is_same_col and not has_values:
            section_had_accounts = False
            continue

        # ── Yardi-style subtotals: have an account code AND "total" in name ──
        # In Yardi/MRI, subtotal rows carry their own account codes (col A ≠ col B)
        # but the description says "Total X" — these always duplicate individual items.
        is_yardi_total = (not is_same_col and has_values and
                          (name_lower.startswith("total ") or
                           name_lower.startswith("subtotal ")))
        if is_yardi_total:
            continue

        # ── CBRE-style "Total X" rows: col A = col B, has values ─────────────
        # Keep if no individual items have been seen yet in this section (leaf data).
        # Skip if individual items exist (duplicate subtotal).
        is_total_row = is_same_col and has_values and (
            name_lower.startswith("total ") or name_lower.startswith("subtotal ")
        )
        if is_total_row:
            if section_had_accounts:
                section_had_accounts = False  # marks section boundary
                continue
            # No individual items → this IS the leaf data; keep it.

        # ── Individual account-code row or blank-acct leaf row ───────────────
        if not is_total_row:
            section_had_accounts = True

        # ── Map to COA and accumulate ─────────────────────────────────────────
        # For name-only format, prefix encodes the section so the same item name
        # in the income and expense sections maps to different COA codes.
        if name_only and not acct_raw:
            base = name_raw.strip().lower()
            prefix = f"{entrata_section}::{base}" if entrata_section != "unknown" else base
        else:
            prefix = _acct_prefix(acct_raw) if acct_raw else name_raw.strip().lower()
        coa_code = acct_map.get(prefix)
        if not coa_code:
            if has_values:
                unmapped.append({"prefix": prefix, "acct": acct_raw, "name": name_raw, "total": total})
            coa_code = "nai" if total >= 0 else "nae"

        line_items.append({
            "acct":      acct_raw,
            "name":      name_raw,
            "coa_code":  coa_code,
            "coa_label": COA_LABELS.get(coa_code, coa_code),
            "monthly":   monthly,
            "total":     total,
        })
        if coa_code not in coa:
            coa[coa_code] = [0.0] * n_months
        for i, v in enumerate(monthly):
            coa[coa_code][i] += v

    return line_items, unmapped, coa, reported_noi


def _excel_to_rows(file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    hdr_row, first_col, total_col, months = _detect_header_row(ws)

    # Name-only format (Entrata/Griffis): col A = description, col B = first month.
    # Detected when first month is in col 2 (1-based), i.e. immediately after col A.
    name_only = (first_col == 2)

    # Trailing reports may have 13+ months; keep only the most recent 12.
    if len(months) > 12:
        skip = len(months) - 12
        first_col += skip   # advance data-start column (1-based)
        months = months[skip:]

    rows = []
    for r in range(hdr_row, ws.max_row + 1):
        rows.append([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])
    # normalize to strings so _parse_rows works uniformly
    str_rows = []
    for row in rows:
        str_rows.append([
            v.strftime("%b %Y") if isinstance(v, datetime) else
            (str(v).strip() if v is not None else "")
            for v in row
        ])
    return str_rows, 0, first_col - 1, total_col - 1, months, name_only


# ── Public API ─────────────────────────────────────────────────────────────────

def parse_t12(file_bytes: bytes, extra_mappings: dict = None) -> dict:
    """
    Parse a T12 operating statement (Excel or PDF).
    extra_mappings: {acct_prefix: coa_code} from Claude classification.
    """
    acct_map = {**_YARDI_ACCT_MAP, **(extra_mappings or {})}
    is_pdf   = file_bytes[:4] == b"%PDF"

    if is_pdf:
        rows = _pdf_to_rows(file_bytes)
        hdr_idx, first_col, total_col, months = _detect_header_row_pdf(rows)
        name_only = False
    else:
        rows, hdr_idx, first_col, total_col, months, name_only = _excel_to_rows(file_bytes)

    n_months = total_col - first_col
    line_items, unmapped, coa_raw, reported_noi = _parse_rows(
        rows, hdr_idx, first_col, total_col, acct_map, name_only
    )

    # ── Sign normalisation ────────────────────────────────────────────────────
    # Some T12 formats (e.g. CBRE Condensed) present income items that offset
    # expenses as negative values in the expense section (utility reimbursements,
    # bad-debt recoveries). Flip any item whose sign contradicts its COA role.
    _must_positive = frozenset({
        "mkt", "aff", "pkg", "stg", "pet", "tv", "oinc", "rubs", "cominc",
        "adv", "adm", "rm", "cs", "to", "pay", "util",
        "mgt", "hoa", "comexp",
        "capx", "nrcapx", "amcapx", "intcapx", "tilc",
        # "ins" and "ret" intentionally excluded: tax credits/abatements and
        # insurance rebates are legitimately negative line items within those categories.
    })
    _must_negative = frozenset({"ltl", "vac", "conc", "empmo", "down", "bd"})

    # "gain" exempts "Gain to Lease" (ltl, legitimately positive) from the must-negative flip.
    _recovery_words = frozenset({"recovery", "recoveries", "reimburse", "reimbursement", "gain"})

    for item in line_items:
        code  = item["coa_code"]
        total = item["total"]
        if abs(total) < 0.01:
            continue
        name_words = set(re.split(r"\W+", item["name"].lower()))
        is_recovery = bool(name_words & _recovery_words)
        needs_flip = (
            (code in _must_positive and total < 0) or
            (code in _must_negative and total > 0 and not is_recovery)
        )
        if needs_flip:
            item["monthly"] = [-v for v in item["monthly"]]
            item["total"]   = -total

    # Rebuild coa_raw from normalised line_items
    coa_raw = {}
    for item in line_items:
        code = item["coa_code"]
        if code not in coa_raw:
            coa_raw[code] = [0.0] * n_months
        for i, v in enumerate(item["monthly"][:n_months]):
            coa_raw[code][i] += v

    result_coa = {
        code: {"label": COA_LABELS.get(code, code), "monthly": monthly, "total": sum(monthly)}
        for code, monthly in coa_raw.items()
    }

    def t(code): return result_coa.get(code, {}).get("total", 0.0)

    gpr          = t("mkt") + t("aff")
    ltl          = t("ltl")
    vac          = t("vac")
    conc         = t("conc")
    empmo        = t("empmo")
    down         = t("down")
    bad_debt     = t("bd")
    net_rental   = gpr + ltl + vac + conc + empmo + down + bad_debt
    other_income = sum(t(c) for c in ("pkg", "stg", "pet", "tv", "oinc", "rubs", "cominc"))
    egi          = net_rental + other_income
    variable_opex    = sum(t(c) for c in ("adv", "adm", "rm", "cs", "to", "pay"))
    nonvariable_opex = sum(t(c) for c in ("util", "mgt", "ins", "ret", "hoa", "comexp"))
    total_opex   = variable_opex + nonvariable_opex
    noi          = egi - total_opex

    def _fmt(v): return f"${abs(v):,.0f}"

    summary = {
        "period":             f"{months[0]} – {months[-1]}" if months else "",
        "gpr":                _fmt(gpr),
        "loss_to_lease":      f"{ltl / gpr * 100:.1f}%" if gpr else None,
        "physical_occupancy": f"{abs(vac) / gpr * 100:.1f}%" if gpr else None,
        "net_rental_income":  _fmt(net_rental),
        "total_other_income": _fmt(other_income),
        "t12_egi":            _fmt(egi),
        "t12_opex":           _fmt(abs(total_opex)),
        "t12_opex_pct":       f"{abs(total_opex) / egi * 100:.1f}%" if egi else None,
        "t12_noi":            _fmt(noi),
        "t12_noi_margin":     f"{noi / egi * 100:.1f}%" if egi else None,
        "in_place_rent":      None,
    }

    return {
        "period":       f"{months[0]} – {months[-1]}" if months else "",
        "months":       months,
        "n_months":     n_months,
        "line_items":   line_items,
        "unmapped":     unmapped,
        "coa":          result_coa,
        "summary":      summary,
        "reported_noi": reported_noi,
        "_gpr": gpr, "_egi": egi, "_total_opex": total_opex, "_noi": noi,
    }
